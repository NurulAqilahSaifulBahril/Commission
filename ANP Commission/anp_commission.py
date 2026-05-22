#!/usr/bin/env python3
"""
ANP commission report for internal and full-time agents.

Data source: prod_main via Postgres SQL proxy API.

Total amount (tier basis):
    invoice total_amount (same as invoice_total_amount in app).

Sales price (informational):
    total_amount - epp_internal

epp_internal:
    total_amount * (effective_epp / 100) when effective_epp > 0, else 0.

Eligibility:
    - Agent agent_type in --agent-types (default: internal + FULL TIME)
    - Invoice has 1st payment secured (1st_payment_date IS NOT NULL)
    - Invoice not soft-deleted (is_deleted IS NOT TRUE)

Commission timing:
    ANP is paid in the calendar month after the invoice_date month,
    once 1st payment is secured (e.g. Jan 2026 invoice -> Feb 2026 payout).
    Each invoicing month accumulates total_amount separately per agent.

ANP tier (on accumulated total amount per agent within each invoicing month):
    RM 0      : below RM 60,000
    RM 500    : RM 60,000  - 179,999.99
    RM 1,000  : RM 180,000 - 359,999.99
    RM 1,500  : RM 360,000 - 719,999.99
    RM 2,000  : RM 720,000 and above
"""

from __future__ import annotations

import argparse
import calendar
import csv
import os
import sys
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

import requests
from dotenv import load_dotenv

load_dotenv()

MONEY = Decimal("0.01")


@dataclass(frozen=True)
class CommissionTier:
    min_inclusive: Decimal
    max_inclusive: Decimal | None
    commission_rm: Decimal


TIERS: tuple[CommissionTier, ...] = (
    CommissionTier(Decimal("0"), Decimal("59999.99"), Decimal("0")),
    CommissionTier(Decimal("60000"), Decimal("179999.99"), Decimal("500")),
    CommissionTier(Decimal("180000"), Decimal("359999.99"), Decimal("1000")),
    CommissionTier(Decimal("360000"), Decimal("719999.99"), Decimal("1500")),
    CommissionTier(Decimal("720000"), None, Decimal("2000")),
)


def normalize_proxy_token(raw: str) -> str:
    """Accept token only, or values pasted as 'Bearer <jwt>' / quoted strings."""
    token = (raw or "").strip().strip('"').strip("'")
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    if token.lower().startswith("authorization:"):
        token = token.split(":", 1)[1].strip()
        if token.lower().startswith("bearer "):
            token = token[7:].strip()
    return token.strip()


def validate_proxy_token(token: str) -> None:
    placeholders = {
        "",
        "your_bearer_token_here",
        "paste_your_jwt_token_here",
        "paste_token_here",
        "changeme",
    }
    if token.lower() in placeholders:
        raise ValueError(
            "PG_PROXY_TOKEN is still the placeholder. Edit .env and paste the JWT "
            "from your Postgres proxy connection packet (token only, no 'Bearer ' prefix)."
        )
    parts = token.split(".")
    if len(parts) != 3:
        raise ValueError(
            "PG_PROXY_TOKEN does not look like a valid JWT (expected 3 parts separated by '.'). "
            "Paste only the token string from the proxy admin, not the full URL or SQL body."
        )


class PostgresProxyClient:
    def __init__(self, base_url: str, token: str, db_name: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.db_name = db_name
        token = normalize_proxy_token(token)
        validate_proxy_token(token)
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            }
        )

    def query(self, sql: str) -> list[dict[str, Any]]:
        response = self.session.post(
            f"{self.base_url}/api/sql",
            json={"db_name": self.db_name, "sql": sql, "params": []},
            timeout=120,
        )
        if not response.ok:
            raise RuntimeError(
                f"SQL proxy error {response.status_code}: {response.text}\nSQL: {sql[:500]}"
            )
        payload = response.json()
        return list(payload.get("rows") or [])


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                continue
    return None


def anp_commission(accumulated_total_amount: Decimal) -> Decimal:
    for tier in TIERS:
        if accumulated_total_amount < tier.min_inclusive:
            continue
        if tier.max_inclusive is None or accumulated_total_amount <= tier.max_inclusive:
            return tier.commission_rm
    return TIERS[-1].commission_rm


def epp_internal_amount(
    total_amount: Decimal, effective_epp: Decimal, payment_1_charges: Decimal | None
) -> Decimal:
    if payment_1_charges and payment_1_charges > 0:
        return payment_1_charges
    if effective_epp > 0:
        return (total_amount * effective_epp / Decimal("100")).quantize(
            MONEY, rounding=ROUND_HALF_UP
        )
    return Decimal("0")


def sales_price(
    total_amount: Decimal, effective_epp: Decimal, payment_1_charges: Decimal | None
) -> Decimal:
    return total_amount - epp_internal_amount(
        total_amount, effective_epp, payment_1_charges
    )


def invoice_period_bounds(payout_year: int, payout_month: int) -> tuple[date, date]:
    """Invoices counted when invoice_date falls in the month before payout month."""
    if payout_month == 1:
        inv_year, inv_month = payout_year - 1, 12
    else:
        inv_year, inv_month = payout_year, payout_month - 1
    start = date(inv_year, inv_month, 1)
    last_day = monthrange(inv_year, inv_month)[1]
    end = date(inv_year, inv_month, last_day)
    return start, end


def invoice_calendar_month_bounds(year: int, month: int) -> tuple[date, date]:
    """First and last day of a calendar invoicing month (invoice_date falls in this range)."""
    start = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end = date(year, month, last_day)
    return start, end


def commission_payout_month_for_invoice_month(inv_year: int, inv_month: int) -> tuple[int, int]:
    """Commission is paid in the month after the invoicing calendar month."""
    if inv_month == 12:
        return inv_year + 1, 1
    return inv_year, inv_month + 1


def anp_commission_payout_ym(inv_date: date | None) -> str:
    """YYYY-MM of the month when ANP is received (month after invoice_date)."""
    if inv_date is None:
        return ""
    py, pm = commission_payout_month_for_invoice_month(inv_date.year, inv_date.month)
    return f"{py:04d}-{pm:02d}"


def anp_commission_date_remark(inv_date: date | None) -> str:
    """Human-readable month when the agent receives ANP for this invoice."""
    if inv_date is None:
        return ""
    py, pm = commission_payout_month_for_invoice_month(inv_date.year, inv_date.month)
    return f"{calendar.month_name[pm]} {py} (paid in this calendar month)"


def count_distinct_agents_in_invoice_year(
    invoices: list[dict[str, Any]],
    allowed_agent_ids: set[str],
    year: int,
) -> int:
    """Agents with at least one qualifying invoice dated in `year`."""
    seen: set[str] = set()
    for inv in invoices:
        if is_deleted_type(inv):
            continue
        inv_d = parse_date(inv.get("invoice_date"))
        if inv_d is None or inv_d.year != year:
            continue
        aid = str(inv.get("linked_agent") or "")
        if aid in allowed_agent_ids:
            seen.add(aid)
    return len(seen)


def sql_in_list(values: Iterable[str]) -> str:
    escaped = ", ".join("'" + v.replace("'", "''") + "'" for v in values)
    return escaped


def fetch_agents(client: PostgresProxyClient, agent_types: list[str]) -> list[dict[str, Any]]:
    types_sql = sql_in_list(agent_types)
    return client.query(
        f"""
        SELECT bubble_id, name, agent_type, unique_id, linked_user_login
        FROM agent
        WHERE agent_type IN ({types_sql})
        ORDER BY name
        """
    )


def fetch_invoices_for_agents(
    client: PostgresProxyClient, agent_bubble_ids: list[str]
) -> list[dict[str, Any]]:
    if not agent_bubble_ids:
        return []
    ids_sql = sql_in_list(agent_bubble_ids)
    return client.query(
        f"""
        SELECT
          bubble_id,
          linked_agent,
          invoice_number,
          invoice_date,
          amount,
          total_amount,
          effective_epp,
          customer_name_snapshot,
          linked_customer,
          "1st_payment_date",
          is_deleted,
          type
        FROM invoice
        WHERE linked_agent IN ({ids_sql})
          AND "1st_payment_date" IS NOT NULL
          AND COALESCE(is_deleted, false) = false
        ORDER BY linked_agent, invoice_date, invoice_number
        """
    )


def fetch_payment_planning(
    client: PostgresProxyClient, invoice_bubble_ids: list[str]
) -> dict[str, dict[str, Any]]:
    if not invoice_bubble_ids:
        return {}
    # Proxy may limit payload size; batch in chunks
    out: dict[str, dict[str, Any]] = {}
    chunk_size = 200
    for i in range(0, len(invoice_bubble_ids), chunk_size):
        chunk = invoice_bubble_ids[i : i + chunk_size]
        ids_sql = sql_in_list(chunk)
        rows = client.query(
            f"""
            SELECT linked_invoice, payment_1_charges
            FROM invoice_payment_planning
            WHERE linked_invoice IN ({ids_sql})
            """
        )
        for row in rows:
            key = row.get("linked_invoice")
            if key:
                out[str(key)] = row
    return out


def fetch_customers(
    client: PostgresProxyClient, customer_ids: list[str]
) -> dict[str, str]:
    if not customer_ids:
        return {}
    out: dict[str, str] = {}
    chunk_size = 200
    for i in range(0, len(customer_ids), chunk_size):
        chunk = customer_ids[i : i + chunk_size]
        ids_sql = sql_in_list(chunk)
        rows = client.query(
            f"""
            SELECT customer_id, name
            FROM customer
            WHERE customer_id IN ({ids_sql})
            """
        )
        for row in rows:
            name = row.get("name") or ""
            if row.get("customer_id"):
                out[str(row["customer_id"])] = name
    return out


def resolve_customer_name(
    invoice: dict[str, Any], customers: dict[str, str]
) -> str:
    snapshot = (invoice.get("customer_name_snapshot") or "").strip()
    if snapshot:
        return snapshot
    linked = invoice.get("linked_customer")
    if linked:
        return customers.get(str(linked), str(linked))
    return ""


def invoice_in_period(
    invoice: dict[str, Any], period_start: date | None, period_end: date | None
) -> bool:
    if period_start is None or period_end is None:
        return True
    inv_date = parse_date(invoice.get("invoice_date"))
    if inv_date is None:
        return False
    return period_start <= inv_date <= period_end


def is_deleted_type(invoice: dict[str, Any]) -> bool:
    inv_type = (invoice.get("type") or "").upper()
    if "DELETED" in inv_type:
        return True
    return False


def build_report_rows(
    agents: list[dict[str, Any]],
    invoices: list[dict[str, Any]],
    planning: dict[str, dict[str, Any]],
    customers: dict[str, str],
    period_start: date | None,
    period_end: date | None,
    payout_label: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    agent_by_id = {str(a["bubble_id"]): a for a in agents}
    detail_rows: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []

    invoices_by_agent: dict[str, list[dict[str, Any]]] = {}
    for inv in invoices:
        if is_deleted_type(inv):
            continue
        if not invoice_in_period(inv, period_start, period_end):
            continue
        agent_id = str(inv.get("linked_agent") or "")
        if agent_id not in agent_by_id:
            continue
        invoices_by_agent.setdefault(agent_id, []).append(inv)

    qualifying_agent_ids = sorted(invoices_by_agent.keys())

    for agent_id in qualifying_agent_ids:
        agent = agent_by_id[agent_id]
        agent_name = (agent.get("name") or "").strip()
        agent_invoices = sorted(
            invoices_by_agent[agent_id],
            key=lambda x: (
                parse_date(x.get("invoice_date")) or date.min,
                str(x.get("invoice_number") or ""),
            ),
        )
        accumulated_total = Decimal("0")
        for inv in agent_invoices:
            inv_date = parse_date(inv.get("invoice_date"))
            total = to_decimal(inv.get("total_amount") or inv.get("amount"))
            epp_pct = to_decimal(inv.get("effective_epp"))
            plan = planning.get(str(inv.get("bubble_id")))
            p1_charges = (
                to_decimal(plan.get("payment_1_charges"))
                if plan and plan.get("payment_1_charges") is not None
                else None
            )
            epp_int = epp_internal_amount(total, epp_pct, p1_charges)
            sp = sales_price(total, epp_pct, p1_charges)
            accumulated_total += total
            detail_rows.append(
                {
                    "payout_period": payout_label,
                    "agent_name": agent_name,
                    "agent_bubble_id": agent_id,
                    "agent_type": agent.get("agent_type"),
                    "customer_name": resolve_customer_name(inv, customers),
                    "invoice_number": inv.get("invoice_number") or inv.get("bubble_id"),
                    "invoice_date": inv_date,
                    "first_payment_date": parse_date(inv.get("1st_payment_date")),
                    "invoice_total_amount": total,
                    "epp_internal": epp_int,
                    "sales_price": sp,
                    "accumulated_total_amount": accumulated_total,
                    "anp_commission_line": Decimal("0"),
                    "anp_commission_accumulated_tier": anp_commission(
                        accumulated_total
                    ),
                    "anp_commission_date": anp_commission_payout_ym(inv_date),
                    "anp_commission_date_remark": anp_commission_date_remark(inv_date),
                }
            )

        final_commission = anp_commission(accumulated_total)
        summary_rows.append(
            {
                "payout_period": payout_label,
                "agent_name": agent_name,
                "agent_bubble_id": agent_id,
                "agent_type": agent.get("agent_type"),
                "invoice_count": len(agent_invoices),
                "accumulated_total_amount": accumulated_total,
                "anp_commission": final_commission,
            }
        )

    return detail_rows, summary_rows


def _fmt_date(val: Any) -> str:
    if isinstance(val, date):
        return val.isoformat()
    return str(val) if val else ""


def _fmt_money(val: Any) -> str:
    if val is None or val == "":
        return "0.00"
    if isinstance(val, Decimal):
        return f"{val:,.2f}"
    return f"{Decimal(str(val)):,.2f}"


def summary_table_rows(summary_rows: list[dict[str, Any]]) -> list[list[str]]:
    """Table 1: agent totals."""
    rows = []
    for r in summary_rows:
        rows.append(
            [
                r.get("agent_name", ""),
                str(r.get("invoice_count", 0)),
                _fmt_money(r.get("accumulated_total_amount")),
                _fmt_money(r.get("anp_commission")),
            ]
        )
    return rows


def detail_table_rows(detail_rows: list[dict[str, Any]]) -> list[list[str]]:
    """Table 2: per-invoice lines."""
    rows = []
    for r in detail_rows:
        rows.append(
            [
                r.get("agent_name", ""),
                r.get("customer_name", ""),
                str(r.get("invoice_number", "")),
                _fmt_date(r.get("invoice_date")),
                _fmt_money(r.get("invoice_total_amount")),
                _fmt_money(r.get("sales_price")),
                _fmt_money(r.get("accumulated_total_amount")),
                _fmt_money(r.get("anp_commission_accumulated_tier")),
                r.get("anp_commission_date", ""),
                (r.get("anp_commission_date_remark") or "")[:60],
            ]
        )
    return rows


def print_report_tables(
    meta: dict[str, Any],
    summary_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    max_detail_console_rows: int | None,
) -> None:
    from tabulate import tabulate

    # ASCII table format works on Windows cp1252 consoles
    table_fmt = "simple"

    print()
    print("=" * 100)
    print("ANP COMMISSION REPORT")
    print("=" * 100)
    overview = [
        ["Billing period label", meta.get("payout_period", "")],
        ["Invoice calendar month", meta.get("invoice_calendar_month", "")],
        ["Commission payout month", meta.get("commission_payout_month", "")],
        [
            "Distinct agents (invoice year)",
            meta.get("distinct_agents_invoice_year", ""),
        ],
        ["Invoice date from", meta.get("invoice_date_from", "")],
        ["Invoice date to", meta.get("invoice_date_to", "")],
        ["Agent types", meta.get("agent_types", "")],
        ["Agents in this period slice", meta.get("total_qualifying_agents", 0)],
        ["Invoices in this period slice", meta.get("total_qualifying_invoices", 0)],
    ]
    print(tabulate(overview, headers=["Field", "Value"], tablefmt=table_fmt, disable_numparse=True))
    print()

    print("TABLE 1 - Agent summary (accumulated total amount & ANP commission)")
    print("-" * 100)
    summary_headers = [
        "Agent Name",
        "Invoices",
        "Accumulated Total (RM)",
        "ANP Commission (RM)",
    ]
    summary_data = summary_table_rows(summary_rows)
    if summary_data:
        print(tabulate(summary_data, headers=summary_headers, tablefmt=table_fmt, disable_numparse=True))
        total_anp = sum(Decimal(str(r.get("anp_commission", 0))) for r in summary_rows)
        total_inv = sum(int(r.get("invoice_count", 0)) for r in summary_rows)
        print(
            tabulate(
                [["TOTAL", str(total_inv), "-", _fmt_money(total_anp)]],
                headers=summary_headers,
                tablefmt=table_fmt,
                disable_numparse=True,
            )
        )
    else:
        print("(no qualifying agents)")
    print()

    print("TABLE 2 - Invoice detail (per customer)")
    print("-" * 100)
    detail_headers = [
        "Agent Name",
        "Customer Name",
        "Invoice #",
        "Invoice Date",
        "Invoice Total (RM)",
        "Sales Price (RM)",
        "Accumulated Total (RM)",
        "ANP Tier (RM)",
        "ANP Date YYYY-MM",
        "ANP month (remark)",
    ]
    detail_data = detail_table_rows(detail_rows)
    if detail_data:
        limit = max_detail_console_rows
        if limit is not None and len(detail_data) > limit:
            print(
                f"(Showing first {limit} of {len(detail_data)} invoice rows in console. "
                "Full table is in the Excel/CSV files.)"
            )
            detail_data = detail_data[:limit]
        print(tabulate(detail_data, headers=detail_headers, tablefmt=table_fmt, disable_numparse=True))
    else:
        print("(no qualifying invoices)")
    print()


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            out = {}
            for key in fieldnames:
                val = row.get(key)
                if isinstance(val, (date, datetime)):
                    out[key] = val.isoformat()
                elif isinstance(val, Decimal):
                    out[key] = f"{val:.2f}"
                else:
                    out[key] = val
            writer.writerow(out)


def _style_sheet_header(
    ws,
    headers: list[str],
    money_cols: set[int] | None = None,
    date_cols: set[int] | None = None,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    money_cols = money_cols or set()
    date_cols = date_cols or set()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for col_idx, cell in enumerate(row, start=1):
            cell.border = border
            if col_idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in date_cols and cell.value:
                cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_excel(
    path: Path,
    summary: list[dict[str, Any]],
    detail: list[dict[str, Any]],
    meta: dict[str, Any],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Overview"
    ws_meta["A1"] = "ANP Commission Report"
    ws_meta["A1"].font = Font(bold=True, size=14)
    row = 3
    for key, value in meta.items():
        ws_meta.cell(row=row, column=1, value=key)
        ws_meta.cell(row=row, column=2, value=value)
        row += 1
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 40

    # Table 1 — Agent summary
    ws_agents = wb.create_sheet("Table1_Agent_Summary")
    agent_headers = [
        "Agent Name",
        "Agent Type",
        "Invoice Count",
        "Accumulated Total (RM)",
        "ANP Commission (RM)",
    ]
    ws_agents.append(agent_headers)
    for r in summary:
        ws_agents.append(
            [
                r.get("agent_name"),
                r.get("agent_type"),
                r.get("invoice_count"),
                float(r.get("accumulated_total_amount", 0)),
                float(r.get("anp_commission", 0)),
            ]
        )
    if summary:
        ws_agents.append(
            [
                "TOTAL",
                "",
                sum(int(r.get("invoice_count", 0)) for r in summary),
                "N/A",
                float(sum(Decimal(str(r.get("anp_commission", 0))) for r in summary)),
            ]
        )
    _style_sheet_header(ws_agents, agent_headers, money_cols={4, 5})

    # Table 2 — Invoice detail
    ws_detail = wb.create_sheet("Table2_Invoice_Detail")
    detail_headers = [
        "Agent Name",
        "Customer Name",
        "Invoice #",
        "Invoice Date",
        "1st Payment Date",
        "Invoice Total (RM)",
        "EPP Internal (RM)",
        "Sales Price (RM)",
        "Accumulated Total (RM)",
        "ANP Tier (RM)",
        "ANP Commission Date YYYY-MM",
        "ANP Commission Month (remark)",
    ]
    ws_detail.append(detail_headers)
    for r in detail:
        ws_detail.append(
            [
                r.get("agent_name"),
                r.get("customer_name"),
                r.get("invoice_number"),
                _fmt_date(r.get("invoice_date")),
                _fmt_date(r.get("first_payment_date")),
                float(r.get("invoice_total_amount", 0)),
                float(r.get("epp_internal", 0)),
                float(r.get("sales_price", 0)),
                float(r.get("accumulated_total_amount", 0)),
                float(r.get("anp_commission_accumulated_tier", 0)),
                r.get("anp_commission_date") or "",
                r.get("anp_commission_date_remark") or "",
            ]
        )
    _style_sheet_header(
        ws_detail, detail_headers, money_cols={6, 7, 8, 9, 10}, date_cols={4, 5}
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


DETAIL_CSV_FIELDS = [
    "payout_period",
    "agent_name",
    "agent_type",
    "customer_name",
    "invoice_number",
    "invoice_date",
    "first_payment_date",
    "invoice_total_amount",
    "epp_internal",
    "sales_price",
    "accumulated_total_amount",
    "anp_commission_line",
    "anp_commission_accumulated_tier",
    "anp_commission_date",
    "anp_commission_date_remark",
]
SUMMARY_CSV_FIELDS = [
    "payout_period",
    "agent_name",
    "agent_type",
    "invoice_count",
    "accumulated_total_amount",
    "anp_commission",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate ANP commission report.")
    parser.add_argument(
        "--payout-month",
        help="Commission payout month (YYYY-MM). Invoices from prior calendar month are included.",
    )
    parser.add_argument(
        "--all-time",
        action="store_true",
        help="Include all qualifying invoices (no invoice_date month filter).",
    )
    parser.add_argument(
        "--agent-types",
        default="internal,FULL TIME",
        help='Comma-separated agent_type values (default: "internal,FULL TIME").',
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Directory for CSV/XLSX output (default: output)",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel workbook generation.",
    )
    parser.add_argument(
        "--full-console",
        action="store_true",
        help="Print every invoice row in the console (default: first 40 rows only).",
    )
    parser.add_argument(
        "--invoice-year",
        type=int,
        metavar="YYYY",
        help="Same as --year-invoice-months: one bundle per invoicing month in that year.",
    )
    parser.add_argument(
        "--headcount-year",
        type=int,
        default=2026,
        metavar="YYYY",
        help=(
            "Print/count distinct agents with at least one qualifying invoice dated in "
            "this year (default: 2026)."
        ),
    )
    parser.add_argument(
        "--year-invoice-months",
        type=int,
        metavar="YYYY",
        help=(
            "Generate one report per calendar invoicing month in that year that has "
            "qualifying data (e.g. 2026 outputs each calendar month Jan-Dec where data exists). "
            "Skips months with no matching invoices."
        ),
    )
    parser.add_argument(
        "--print-each-month",
        action="store_true",
        help="With --year-invoice-months: print full console tables for every non-empty month.",
    )
    parser.add_argument(
        "--console-detail-limit",
        type=int,
        default=40,
        metavar="N",
        help="Max invoice rows in console table (default: 40). Use with --full-console for all rows.",
    )
    return parser.parse_args()


def print_year_month_overview(rows: list[list[Any]]) -> None:
    from tabulate import tabulate

    headers = [
        "Invoice month",
        "Commission payout",
        "Agents",
        "Invoices",
        "Total ANP (RM)",
    ]
    print(tabulate(rows, headers=headers, tablefmt="simple", disable_numparse=True))


def main() -> int:
    args = parse_args()

    if args.invoice_year is not None and args.year_invoice_months is not None:
        print(
            "Use only one of --invoice-year and --year-invoice-months.",
            file=sys.stderr,
        )
        return 1
    if args.invoice_year is not None:
        args.year_invoice_months = args.invoice_year

    base_url = os.getenv("PG_PROXY_URL", "").strip()
    token = normalize_proxy_token(os.getenv("PG_PROXY_TOKEN", ""))
    db_name = os.getenv("PG_DB_NAME", "prod_main").strip()

    if not base_url or not token:
        print(
            "Set PG_PROXY_URL and PG_PROXY_TOKEN in .env (see .env.example).",
            file=sys.stderr,
        )
        return 1

    try:
        validate_proxy_token(token)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print(
            f"Edit this file: {Path('.env').resolve()}",
            file=sys.stderr,
        )
        return 1

    mode_flags = sum(
        [
            bool(args.all_time),
            bool(args.payout_month),
            args.year_invoice_months is not None,
        ]
    )
    if mode_flags > 1:
        print(
            "Use only one period mode: --all-time, --payout-month YYYY-MM, or --year-invoice-months YYYY",
            file=sys.stderr,
        )
        return 1

    agent_types = [t.strip() for t in args.agent_types.split(",") if t.strip()]
    if not agent_types:
        print("At least one --agent-types value is required.", file=sys.stderr)
        return 1

    client = PostgresProxyClient(base_url, token, db_name)

    print("Fetching agents...")
    agents = fetch_agents(client, agent_types)
    agent_ids = [str(a["bubble_id"]) for a in agents]
    print(f"  Agents with types {agent_types}: {len(agents)}")

    print("Fetching invoices (1st payment secured)...")
    invoices = fetch_invoices_for_agents(client, agent_ids)
    print(f"  Raw invoices: {len(invoices)}")

    invoice_ids = [str(i["bubble_id"]) for i in invoices if i.get("bubble_id")]
    customer_ids = list(
        {str(i["linked_customer"]) for i in invoices if i.get("linked_customer")}
    )

    print("Fetching payment planning (EPP charges)...")
    planning = fetch_payment_planning(client, invoice_ids)

    print("Fetching customer names...")
    customers = fetch_customers(client, customer_ids)

    agent_id_set = set(agent_ids)
    n_distinct_headcount = count_distinct_agents_in_invoice_year(
        invoices, agent_id_set, args.headcount_year
    )
    print(
        f"\nDistinct agents with >=1 qualifying invoice dated in "
        f"{args.headcount_year}: {n_distinct_headcount}\n"
    )
    headcount_meta = (
        f"{n_distinct_headcount} (invoice_date in {args.headcount_year})"
    )

    out_dir = Path(args.output_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    agent_types_str = ", ".join(agent_types)

    if args.year_invoice_months:
        year = args.year_invoice_months
        overview_rows: list[list[Any]] = []

        for inv_month in range(1, 13):
            ps, pe = invoice_calendar_month_bounds(year, inv_month)
            pay_y, pay_m = commission_payout_month_for_invoice_month(year, inv_month)
            label = payout_label_calendar_invoice_month(year, inv_month)
            calendar_label = f"{year:04d}-{inv_month:02d}"
            pay_label = f"{pay_y:04d}-{pay_m:02d}"

            detail_rows, summary_rows = build_report_rows(
                agents,
                invoices,
                planning,
                customers,
                ps,
                pe,
                label,
            )
            if not detail_rows:
                continue

            total_anp = sum(
                Decimal(str(r.get("anp_commission", 0))) for r in summary_rows
            )
            overview_rows.append(
                [
                    calendar_label,
                    pay_label,
                    len(summary_rows),
                    len(detail_rows),
                    _fmt_money(total_anp),
                ]
            )

            prefix = f"anp_commission_{calendar_label}_{stamp}"
            meta = {
                "payout_period": label,
                "invoice_calendar_month": calendar_label,
                "commission_payout_month": pay_label,
                "invoice_date_from": ps.isoformat(),
                "invoice_date_to": pe.isoformat(),
                "distinct_agents_invoice_year": headcount_meta,
                "agent_types": agent_types_str,
                "total_qualifying_agents": len(summary_rows),
                "total_qualifying_invoices": len(detail_rows),
                "note_agent_filter": (
                    "Default agent types: internal + FULL TIME; override with --agent-types."
                ),
                "accumulated_basis": "Sum of invoice total_amount within each invoicing month",
                "sales_price_formula": "total_amount - epp_internal (informational)",
                "epp_internal_formula": (
                    "total_amount * (effective_epp/100) or payment_1_charges"
                ),
            }

            dc = out_dir / f"{prefix}_detail.csv"
            sc = out_dir / f"{prefix}_agent_summary.csv"
            write_csv(dc, detail_rows, DETAIL_CSV_FIELDS)
            write_csv(sc, summary_rows, SUMMARY_CSV_FIELDS)

            xlsx_path = out_dir / f"{prefix}.xlsx"
            if not args.no_excel:
                try:
                    write_excel(xlsx_path, summary_rows, detail_rows, meta)
                except ImportError:
                    print("openpyxl not installed; skipped Excel.", file=sys.stderr)

            if args.print_each_month:
                print(f"\n--- {calendar_label} (paid {pay_label}) ---")
                try:
                    detail_limit = (
                        None if args.full_console else max(0, args.console_detail_limit)
                    )
                    print_report_tables(meta, summary_rows, detail_rows, detail_limit)
                except ImportError:
                    print(
                        "Install tabulate for table output: pip install tabulate",
                        file=sys.stderr,
                    )

        if not overview_rows:
            print(
                f"No qualifying invoices found for invoicing months in {year} "
                "(check agent types and filters).",
                file=sys.stderr,
            )
            return 0

        print()
        print("=" * 100)
        print(f"ANP COMMISSION - ALL INVOICE MONTHS IN {year}")
        print("=" * 100)
        print(
            "One report per invoicing calendar month where data exists "
            "(first payment secured). ANP is paid in the month after invoice_date."
        )
        print()
        try:
            print_year_month_overview(overview_rows)
        except ImportError:
            for row in overview_rows:
                print("  ", row)
        print()

        overview_csv = (
            out_dir / f"anp_commission_{year}_months_overview_{stamp}.csv"
        )
        overview_csv.parent.mkdir(parents=True, exist_ok=True)
        with overview_csv.open("w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerow(
                [
                    "invoice_month",
                    "commission_payout_month",
                    "agent_count",
                    "invoice_count",
                    "total_anp_rm",
                ]
            )
            for row in overview_rows:
                writer.writerow(row)

        print(f"Month overview CSV: {overview_csv.resolve()}")
        print(f"Generated {len(overview_rows)} monthly report bundles under: {out_dir.resolve()}")

        return 0

    period_start: date | None = None
    period_end: date | None = None

    if args.all_time:
        payout_label = "all-time"
        invoice_calendar_month = "all"
        commission_payout_month = "N/A"
        if args.payout_month:
            print(
                "Warning: --payout-month ignored when --all-time is set.", file=sys.stderr
            )
    else:
        if args.payout_month:
            py, pm = map(int, args.payout_month.split("-"))
        else:
            tdy = date.today()
            py, pm = tdy.year, tdy.month
        period_start, period_end = invoice_period_bounds(py, pm)
        invoice_calendar_month = (
            f"{period_start.year:04d}-{period_start.month:02d}"
            if period_start
            else "unknown"
        )
        commission_payout_month = f"{py:04d}-{pm:02d}"
        payout_label = f"inv-{invoice_calendar_month}_pay-{commission_payout_month}"

    detail_rows, summary_rows = build_report_rows(
        agents,
        invoices,
        planning,
        customers,
        period_start,
        period_end,
        payout_label,
    )

    total_qualifying_agents = len(summary_rows)
    total_invoices = len(detail_rows)

    prefix = f"anp_commission_{payout_label}_{stamp}"

    detail_csv = out_dir / f"{prefix}_detail.csv"
    summary_csv = out_dir / f"{prefix}_agent_summary.csv"
    write_csv(detail_csv, detail_rows, DETAIL_CSV_FIELDS)
    write_csv(summary_csv, summary_rows, SUMMARY_CSV_FIELDS)

    meta = {
        "payout_period": payout_label,
        "invoice_calendar_month": invoice_calendar_month,
        "commission_payout_month": commission_payout_month,
        "invoice_date_from": period_start.isoformat() if period_start else "all",
        "invoice_date_to": period_end.isoformat() if period_end else "all",
        "distinct_agents_invoice_year": headcount_meta,
        "agent_types": agent_types_str,
        "total_qualifying_agents": total_qualifying_agents,
        "total_qualifying_invoices": total_invoices,
        "note_agent_filter": (
            "Default agent types: internal + FULL TIME; override with --agent-types."
        ),
        "accumulated_basis": "Sum of invoice total_amount in this period slice",
        "sales_price_formula": "total_amount - epp_internal (informational)",
        "epp_internal_formula": "total_amount * (effective_epp/100) or payment_1_charges",
    }

    xlsx_path = out_dir / f"{prefix}.xlsx"
    if not args.no_excel:
        try:
            write_excel(xlsx_path, summary_rows, detail_rows, meta)
        except ImportError:
            print("openpyxl not installed; skipped Excel. CSV files written.", file=sys.stderr)

    try:
        detail_limit = None if args.full_console else max(0, args.console_detail_limit)
        print_report_tables(meta, summary_rows, detail_rows, detail_limit)
    except ImportError:
        print("Install tabulate for table output: pip install tabulate", file=sys.stderr)

    print("Files saved:")
    print(f"  Excel (tables): {xlsx_path.resolve() if xlsx_path.exists() else '(skipped)'}")
    print(f"  Summary CSV:    {summary_csv.resolve()}")
    print(f"  Detail CSV:     {detail_csv.resolve()}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
