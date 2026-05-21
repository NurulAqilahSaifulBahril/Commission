#!/usr/bin/env python3
"""
ANP commission report for internal full-time agents.

Data source: prod_main via Postgres SQL proxy API.

Amount per invoice: invoice total_amount (EPP not applied).

Eligibility:
    - Agent agent_type = 'internal' (see --agent-types)
    - Invoice has 1st payment secured (1st_payment_date IS NOT NULL)
    - Invoice not soft-deleted (is_deleted IS NOT TRUE)

Commission timing:
    Payout in month M includes invoices with invoice_date in month M-1
    (e.g. January invoice -> February commission), once 1st payment is secured.

ANP tier (on accumulated total amount per agent within each invoicing month):
    RM 0      : below RM 60,000
    RM 500    : RM 60,000  - 179,999.99
    RM 1,000  : RM 180,000 - 359,999.99
    RM 1,500  : RM 360,000 - 719,999.99
    RM 2,000  : RM 720,000 and above
"""

from __future__ import annotations

import argparse
import calendar
import csv
import os
import sys
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

import requests
from dotenv import load_dotenv

load_dotenv()

MONEY = Decimal("0.01")


@dataclass(frozen=True)
class CommissionTier:
    min_inclusive: Decimal
    max_inclusive: Decimal | None
    commission_rm: Decimal


TIERS: tuple[CommissionTier, ...] = (
    CommissionTier(Decimal("0"), Decimal("59999.99"), Decimal("0")),
    CommissionTier(Decimal("60000"), Decimal("179999.99"), Decimal("500")),
    CommissionTier(Decimal("180000"), Decimal("359999.99"), Decimal("1000")),
    CommissionTier(Decimal("360000"), Decimal("719999.99"), Decimal("1500")),
    CommissionTier(Decimal("720000"), None, Decimal("2000")),
)


def normalize_proxy_token(raw: str) -> str:
    """Accept token only, or values pasted as 'Bearer <jwt>' / quoted strings."""
    token = (raw or "").strip().strip('"').strip("'")
    if token.lower().startswith("bearer "):
        token = token[7:].strip()
    if token.lower().startswith("authorization:"):
        token = token.split(":", 1)[1].strip()
        if token.lower().startswith("bearer "):
            token = token[7:].strip()
    return token.strip()


def validate_proxy_token(token: str) -> None:
    placeholders = {
        "",
        "your_bearer_token_here",
        "paste_your_jwt_token_here",
        "paste_token_here",
        "changeme",
    }
    if token.lower() in placeholders:
        raise ValueError(
            "PG_PROXY_TOKEN is still the placeholder. Edit .env and paste the JWT "
            "from your Postgres proxy connection packet (token only, no 'Bearer ' prefix)."
        )
    parts = token.split(".")
    if len(parts) != 3:
        raise ValueError(
            "PG_PROXY_TOKEN does not look like a valid JWT (expected 3 parts separated by '.'). "
            "Paste only the token string from the proxy admin, not the full URL or SQL body."
        )


class PostgresProxyClient:
    def __init__(self, base_url: str, token: str, db_name: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.db_name = db_name
        token = normalize_proxy_token(token)
        validate_proxy_token(token)
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            }
        )

    def query(self, sql: str) -> list[dict[str, Any]]:
        response = self.session.post(
            f"{self.base_url}/api/sql",
            json={"db_name": self.db_name, "sql": sql, "params": []},
            timeout=120,
        )
        if not response.ok:
            raise RuntimeError(
                f"SQL proxy error {response.status_code}: {response.text}\nSQL: {sql[:500]}"
            )
        payload = response.json()
        return list(payload.get("rows") or [])


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    text = str(value).strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                continue
    return None


def anp_commission(accumulated_total_amount: Decimal) -> Decimal:
    for tier in TIERS:
        if accumulated_total_amount < tier.min_inclusive:
            continue
        if tier.max_inclusive is None or accumulated_total_amount <= tier.max_inclusive:
            return tier.commission_rm
    return TIERS[-1].commission_rm


def anp_commission_payout_ym(inv_date: date | None) -> str:
    if inv_date is None:
        return ""
    py, pm = commission_payout_month_for_invoice_month(inv_date.year, inv_date.month)
    return f"{py:04d}-{pm:02d}"


def count_distinct_agents_in_year(
    invoices: list[dict[str, Any]],
    allowed_agent_ids: set[str],
    year: int,
) -> int:
    seen: set[str] = set()
    for inv in invoices:
        if is_deleted_type(inv):
            continue
        inv_date = parse_date(inv.get("invoice_date"))
        if inv_date is None or inv_date.year != year:
            continue
        aid = str(inv.get("linked_agent") or "")
        if aid in allowed_agent_ids:
            seen.add(aid)
    return len(seen)


def invoice_period_bounds(payout_year: int, payout_month: int) -> tuple[date, date]:
    """Invoices counted when invoice_date falls in the month before payout month."""
    if payout_month == 1:
        inv_year, inv_month = payout_year - 1, 12
    else:
        inv_year, inv_month = payout_year, payout_month - 1
    start = date(inv_year, inv_month, 1)
    last_day = monthrange(inv_year, inv_month)[1]
    end = date(inv_year, inv_month, last_day)
    return start, end


def invoice_calendar_month_bounds(year: int, month: int) -> tuple[date, date]:
    """First and last day of a calendar invoicing month (invoice_date falls in this range)."""
    start = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end = date(year, month, last_day)
    return start, end


def commission_payout_month_for_invoice_month(inv_year: int, inv_month: int) -> tuple[int, int]:
    """Commission is paid in the month after the invoicing calendar month."""
    if inv_month == 12:
        return inv_year + 1, 1
    return inv_year, inv_month + 1


def anp_month_receive(inv_date: date | None) -> str:
    """Human-readable month the agent receives ANP for this invoice."""
    if inv_date is None:
        return ""
    py, pm = commission_payout_month_for_invoice_month(inv_date.year, inv_date.month)
    return f"{calendar.month_name[pm]} {py}"


def payout_label_calendar_invoice_month(inv_year: int, inv_month: int) -> str:
    py, pm = commission_payout_month_for_invoice_month(inv_year, inv_month)
    return f"inv-{inv_year:04d}-{inv_month:02d}_pay-{py:04d}-{pm:02d}"


def sql_in_list(values: Iterable[str]) -> str:
    escaped = ", ".join("'" + v.replace("'", "''") + "'" for v in values)
    return escaped


def fetch_agents(client: PostgresProxyClient, agent_types: list[str]) -> list[dict[str, Any]]:
    types_sql = sql_in_list(agent_types)
    return client.query(
        f"""
        SELECT bubble_id, name, agent_type, unique_id, linked_user_login
        FROM agent
        WHERE agent_type IN ({types_sql})
        ORDER BY name
        """
    )


def fetch_invoices_for_agents(
    client: PostgresProxyClient, agent_bubble_ids: list[str]
) -> list[dict[str, Any]]:
    if not agent_bubble_ids:
        return []
    ids_sql = sql_in_list(agent_bubble_ids)
    return client.query(
        f"""
        SELECT
          bubble_id,
          linked_agent,
          invoice_number,
          invoice_date,
          amount,
          total_amount,
          customer_name_snapshot,
          linked_customer,
          "1st_payment_date",
          is_deleted,
          type
        FROM invoice
        WHERE linked_agent IN ({ids_sql})
          AND "1st_payment_date" IS NOT NULL
          AND COALESCE(is_deleted, false) = false
        ORDER BY linked_agent, invoice_date, invoice_number
        """
    )


def fetch_customers(
    client: PostgresProxyClient, customer_ids: list[str]
) -> dict[str, str]:
    if not customer_ids:
        return {}
    out: dict[str, str] = {}
    chunk_size = 200
    for i in range(0, len(customer_ids), chunk_size):
        chunk = customer_ids[i : i + chunk_size]
        ids_sql = sql_in_list(chunk)
        rows = client.query(
            f"""
            SELECT customer_id, name
            FROM customer
            WHERE customer_id IN ({ids_sql})
            """
        )
        for row in rows:
            name = row.get("name") or ""
            if row.get("customer_id"):
                out[str(row["customer_id"])] = name
    return out


def resolve_customer_name(
    invoice: dict[str, Any], customers: dict[str, str]
) -> str:
    snapshot = (invoice.get("customer_name_snapshot") or "").strip()
    if snapshot:
        return snapshot
    linked = invoice.get("linked_customer")
    if linked:
        return customers.get(str(linked), str(linked))
    return ""


def invoice_in_period(
    invoice: dict[str, Any], period_start: date | None, period_end: date | None
) -> bool:
    if period_start is None or period_end is None:
        return True
    inv_date = parse_date(invoice.get("invoice_date"))
    if inv_date is None:
        return False
    return period_start <= inv_date <= period_end


def is_deleted_type(invoice: dict[str, Any]) -> bool:
    inv_type = (invoice.get("type") or "").upper()
    if "DELETED" in inv_type:
        return True
    return False


def count_distinct_agents_in_year(
    invoices: list[dict[str, Any]],
    allowed_agent_ids: set[str],
    year: int,
) -> int:
    seen: set[str] = set()
    for inv in invoices:
        if is_deleted_type(inv):
            continue
        inv_date = parse_date(inv.get("invoice_date"))
        if inv_date is None or inv_date.year != year:
            continue
        aid = str(inv.get("linked_agent") or "")
        if aid in allowed_agent_ids:
            seen.add(aid)
    return len(seen)


def table2_sort_key(row: dict[str, Any]) -> tuple:
    inv_date = row.get("invoice_date")
    if not isinstance(inv_date, date):
        inv_date = parse_date(inv_date) or date.min
    return (
        row.get("agent_name", ""),
        inv_date,
        str(row.get("invoice_number", "")),
    )


def table1_sort_key(row: dict[str, Any]) -> tuple:
    return (
        row.get("agent_name", ""),
        row.get("invoice_year", 0),
        row.get("invoice_month", 0),
    )


def finalize_table2_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    Recompute running accumulated total amount and ANP commission per agent
    per calendar invoice month (matches Table 1 monthly tier logic).
    """
    sorted_rows = sorted(rows, key=table2_sort_key)
    out: list[dict[str, Any]] = []
    current_month_key: tuple[str, int, int] | None = None
    running = Decimal("0")

    for row in sorted_rows:
        inv_date = row.get("invoice_date")
        if not isinstance(inv_date, date):
            inv_date = parse_date(inv_date) or date.min
        month_key = (
            row.get("agent_name", ""),
            inv_date.year,
            inv_date.month,
        )
        if month_key != current_month_key:
            current_month_key = month_key
            running = Decimal("0")

        line_total = to_decimal(row.get("total_amount"))
        running += line_total
        out.append(
            {
                **row,
                "accumulated_total_amount": running,
                "anp_commission": anp_commission(running),
            }
        )
    return out


def build_anp_tables(
    agents: list[dict[str, Any]],
    invoices: list[dict[str, Any]],
    customers: dict[str, str],
    period_start: date | None,
    period_end: date | None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Build Table 1 (per agent per invoicing month) and Table 2 (per invoice).

    Tier uses accumulated total_amount within each (agent, year, month) group.
    ANP Commission Date = month after invoice_date (Jan invoice -> Feb payout).
    """
    agent_by_id = {str(a["bubble_id"]): a for a in agents}
    table1_rows: list[dict[str, Any]] = []
    table2_rows: list[dict[str, Any]] = []

    # (agent_id, inv_year, inv_month) -> list of invoice dicts
    buckets: dict[tuple[str, int, int], list[dict[str, Any]]] = {}

    for inv in invoices:
        if is_deleted_type(inv):
            continue
        if not invoice_in_period(inv, period_start, period_end):
            continue
        inv_date = parse_date(inv.get("invoice_date"))
        if inv_date is None:
            continue
        agent_id = str(inv.get("linked_agent") or "")
        if agent_id not in agent_by_id:
            continue
        key = (agent_id, inv_date.year, inv_date.month)
        buckets.setdefault(key, []).append(inv)

    for agent_id, inv_year, inv_month in sorted(
        buckets.keys(), key=lambda k: (k[0], k[1], k[2])
    ):
        inv_list = buckets[(agent_id, inv_year, inv_month)]
        agent = agent_by_id[agent_id]
        agent_name = (agent.get("name") or "").strip()
        month_total = Decimal("0")
        for inv in sorted(
            inv_list,
            key=lambda x: (
                parse_date(x.get("invoice_date")) or date.min,
                str(x.get("invoice_number") or ""),
            ),
        ):
            month_total += to_decimal(inv.get("total_amount") or inv.get("amount"))

        commission_rm = anp_commission(month_total)
        pay_y, pay_m = commission_payout_month_for_invoice_month(inv_year, inv_month)
        payout_ym = f"{pay_y:04d}-{pay_m:02d}"

        table1_rows.append(
            {
                "agent_name": agent_name,
                "invoice_year": inv_year,
                "invoice_month": inv_month,
                "accumulated_total_amount": month_total,
                "anp_commission": commission_rm,
                "anp_commission_date": payout_ym,
            }
        )

        for inv in sorted(
            inv_list,
            key=lambda x: (
                parse_date(x.get("invoice_date")) or date.min,
                str(x.get("invoice_number") or ""),
            ),
        ):
            inv_date = parse_date(inv.get("invoice_date"))
            total = to_decimal(inv.get("total_amount") or inv.get("amount"))
            payout_ym = anp_commission_payout_ym(inv_date) if inv_date else ""

            table2_rows.append(
                {
                    "agent_name": agent_name,
                    "customer_name": resolve_customer_name(inv, customers),
                    "invoice_number": inv.get("invoice_number") or inv.get("bubble_id"),
                    "invoice_date": inv_date,
                    "total_amount": total,
                    "anp_commission_date": payout_ym,
                }
            )

    table2_rows = finalize_table2_rows(table2_rows)
    return table1_rows, table2_rows


def _fmt_date(val: Any) -> str:
    if isinstance(val, date):
        return val.isoformat()
    return str(val) if val else ""


def _fmt_money(val: Any) -> str:
    if val is None or val == "":
        return "0.00"
    if isinstance(val, Decimal):
        return f"{val:,.2f}"
    return f"{Decimal(str(val)):,.2f}"


TABLE1_HEADERS = [
    "Agent Name",
    "Accumulated Total Amount (RM)",
    "ANP Commission (RM)",
    "ANP Commission Date",
]

TABLE2_HEADERS = [
    "Agent Name",
    "Customer Name",
    "Invoice #",
    "Invoice Date",
    "Total Amount (RM)",
    "Accumulated Total Amount (RM)",
    "ANP Commission (RM)",
    "ANP Commission Date",
]


def table1_display_rows(rows: list[dict[str, Any]]) -> list[list[str]]:
    out = []
    for r in sorted(rows, key=table1_sort_key):
        out.append(
            [
                r.get("agent_name", ""),
                _fmt_money(r.get("accumulated_total_amount")),
                _fmt_money(r.get("anp_commission")),
                r.get("anp_commission_date", ""),
            ]
        )
    return out


def table2_display_rows(rows: list[dict[str, Any]]) -> list[list[str]]:
    out = []
    for r in sorted(rows, key=table2_sort_key):
        out.append(
            [
                r.get("agent_name", ""),
                r.get("customer_name", ""),
                str(r.get("invoice_number", "")),
                _fmt_date(r.get("invoice_date")),
                _fmt_money(r.get("total_amount")),
                _fmt_money(r.get("accumulated_total_amount")),
                _fmt_money(r.get("anp_commission")),
                r.get("anp_commission_date", ""),
            ]
        )
    return out


def print_report_tables(
    meta: dict[str, Any],
    table1_rows: list[dict[str, Any]],
    table2_rows: list[dict[str, Any]],
    max_detail_console_rows: int | None,
) -> None:
    from tabulate import tabulate

    table_fmt = "simple"

    print()
    print("=" * 100)
    print("ANP COMMISSION REPORT")
    print("=" * 100)
    overview = [
        ["Filter", meta.get("filter_description", "")],
        ["Invoice date from", meta.get("invoice_date_from", "")],
        ["Invoice date to", meta.get("invoice_date_to", "")],
        ["Agent types", meta.get("agent_types", "")],
        ["Total qualifying agents (distinct)", meta.get("total_qualifying_agents", 0)],
        ["Table 1 rows (agent-month)", len(table1_rows)],
        ["Table 2 rows (invoices)", len(table2_rows)],
    ]
    print(tabulate(overview, headers=["Field", "Value"], tablefmt=table_fmt, disable_numparse=True))
    print()

    print("TABLE 1 - Accumulated ANP Commission")
    print("-" * 100)
    t1 = table1_display_rows(table1_rows)
    if t1:
        print(tabulate(t1, headers=TABLE1_HEADERS, tablefmt=table_fmt, disable_numparse=True))
        total_anp = sum(Decimal(str(r.get("anp_commission", 0))) for r in table1_rows)
        print(
            tabulate(
                [["TOTAL", "-", _fmt_money(total_anp), "-"]],
                headers=TABLE1_HEADERS,
                tablefmt=table_fmt,
                disable_numparse=True,
            )
        )
    else:
        print("(no qualifying data for Table 1)")
    print()

    print("TABLE 2 - ANP commission by customer")
    print("-" * 100)
    t2 = table2_display_rows(table2_rows)
    if t2:
        limit = max_detail_console_rows
        if limit is not None and len(t2) > limit:
            print(
                f"(Showing first {limit} of {len(t2)} invoice rows in console. "
                "Full output is in the Excel/CSV files.)"
            )
            t2 = t2[:limit]
        print(tabulate(t2, headers=TABLE2_HEADERS, tablefmt=table_fmt, disable_numparse=True))
    else:
        print("(no qualifying data for Table 2)")
    print()


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            out = {}
            for key in fieldnames:
                val = row.get(key)
                if isinstance(val, (date, datetime)):
                    out[key] = val.isoformat()
                elif isinstance(val, Decimal):
                    out[key] = f"{val:.2f}"
                else:
                    out[key] = val
            writer.writerow(out)


def _style_sheet_header(
    ws,
    headers: list[str],
    money_cols: set[int] | None = None,
    date_cols: set[int] | None = None,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    money_cols = money_cols or set()
    date_cols = date_cols or set()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for col_idx, cell in enumerate(row, start=1):
            cell.border = border
            if col_idx in money_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in date_cols and cell.value:
                cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_excel(
    path: Path,
    table1_rows: list[dict[str, Any]],
    table2_rows: list[dict[str, Any]],
    meta: dict[str, Any],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Overview"
    ws_meta["A1"] = "ANP Commission Report"
    ws_meta["A1"].font = Font(bold=True, size=14)
    row = 3
    for key, value in meta.items():
        ws_meta.cell(row=row, column=1, value=key)
        ws_meta.cell(row=row, column=2, value=value)
        row += 1
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 40

    # Table 1 — Accumulated ANP Commission
    ws1 = wb.create_sheet("Table1_Accumulated_ANP")
    ws1.append(TABLE1_HEADERS)
    for row in table1_display_rows(table1_rows):
        ws1.append(row)
    _style_sheet_header(ws1, TABLE1_HEADERS, money_cols={2, 3}, date_cols={4})

    # Table 2 — ANP commission by customer
    ws2 = wb.create_sheet("Table2_ANP_by_Customer")
    ws2.append(TABLE2_HEADERS)
    for row in table2_display_rows(table2_rows):
        ws2.append(row)
    _style_sheet_header(ws2, TABLE2_HEADERS, money_cols={5, 6, 7}, date_cols={4})

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


TABLE1_CSV_FIELDS = [
    "agent_name",
    "accumulated_total_amount",
    "anp_commission",
    "anp_commission_date",
]
TABLE2_CSV_FIELDS = [
    "agent_name",
    "customer_name",
    "invoice_number",
    "invoice_date",
    "total_amount",
    "accumulated_total_amount",
    "anp_commission",
    "anp_commission_date",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate ANP commission report.")
    parser.add_argument(
        "--payout-month",
        help="Commission payout month (YYYY-MM). Invoices from prior calendar month are included.",
    )
    parser.add_argument(
        "--all-time",
        action="store_true",
        help="Include all qualifying invoices (no invoice_date month filter).",
    )
    parser.add_argument(
        "--agent-types",
        default="internal,FULL TIME",
        help='Comma-separated agent_type values (default: "internal,FULL TIME").',
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Directory for CSV/XLSX output (default: output)",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel workbook generation.",
    )
    parser.add_argument(
        "--full-console",
        action="store_true",
        help="Print every invoice row in the console (default: first 40 rows only).",
    )
    parser.add_argument(
        "--year-invoice-months",
        type=int,
        metavar="YYYY",
        help=(
            "Generate one report per calendar invoicing month in that year that has "
            "qualifying data (e.g. 2026 outputs each calendar month Jan-Dec where data exists). "
            "Skips months with no matching invoices."
        ),
    )
    parser.add_argument(
        "--print-each-month",
        action="store_true",
        help="With --year-invoice-months: print full console tables for every non-empty month.",
    )
    parser.add_argument(
        "--console-detail-limit",
        type=int,
        default=40,
        metavar="N",
        help="Max invoice rows in console table (default: 40). Use with --full-console for all rows.",
    )
    return parser.parse_args()


def print_year_month_overview(rows: list[list[Any]]) -> None:
    from tabulate import tabulate

    headers = [
        "Invoice month",
        "Commission payout",
        "Agents",
        "Invoices",
        "Total ANP (RM)",
    ]
    print(tabulate(rows, headers=headers, tablefmt="simple", disable_numparse=True))


def main() -> int:
    args = parse_args()

    base_url = os.getenv("PG_PROXY_URL", "").strip()
    token = normalize_proxy_token(os.getenv("PG_PROXY_TOKEN", ""))
    db_name = os.getenv("PG_DB_NAME", "prod_main").strip()

    if not base_url or not token:
        print(
            "Set PG_PROXY_URL and PG_PROXY_TOKEN in .env (see .env.example).",
            file=sys.stderr,
        )
        return 1

    try:
        validate_proxy_token(token)
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print(
            f"Edit this file: {Path('.env').resolve()}",
            file=sys.stderr,
        )
        return 1

    mode_flags = sum(
        [
            bool(args.all_time),
            bool(args.payout_month),
            args.year_invoice_months is not None,
        ]
    )
    default_invoice_year_mode = mode_flags == 0
    if mode_flags == 0:
        # Default behavior requested by user: all invoices in 2026 (single report).
        args.year_invoice_months = None
    if mode_flags > 1:
        print(
            "Use only one period mode: --all-time, --payout-month YYYY-MM, or --year-invoice-months YYYY",
            file=sys.stderr,
        )
        return 1

    agent_types = [t.strip() for t in args.agent_types.split(",") if t.strip()]
    if not agent_types:
        print("At least one --agent-types value is required.", file=sys.stderr)
        return 1

    client = PostgresProxyClient(base_url, token, db_name)

    print("Fetching agents...")
    agents = fetch_agents(client, agent_types)
    agent_ids = [str(a["bubble_id"]) for a in agents]
    print(f"  Agents with types {agent_types}: {len(agents)}")

    print("Fetching invoices (1st payment secured)...")
    invoices = fetch_invoices_for_agents(client, agent_ids)
    print(f"  Raw invoices: {len(invoices)}")

    customer_ids = list(
        {str(i["linked_customer"]) for i in invoices if i.get("linked_customer")}
    )

    print("Fetching customer names...")
    customers = fetch_customers(client, customer_ids)

    out_dir = Path(args.output_dir)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    agent_types_str = ", ".join(agent_types)

    if args.year_invoice_months:
        year = args.year_invoice_months
        overview_rows: list[list[Any]] = []

        months = list(range(1, 13))
        for inv_month in months:
            ps, pe = invoice_calendar_month_bounds(year, inv_month)
            pay_y, pay_m = commission_payout_month_for_invoice_month(year, inv_month)
            label = payout_label_calendar_invoice_month(year, inv_month)
            calendar_label = f"{year:04d}-{inv_month:02d}"
            pay_label = f"{pay_y:04d}-{pay_m:02d}"

            table1_rows, table2_rows = build_anp_tables(
                agents, invoices, customers, ps, pe
            )
            if not table2_rows:
                continue

            total_anp = sum(Decimal(str(r.get("anp_commission", 0))) for r in table1_rows)
            overview_rows.append(
                [
                    calendar_label,
                    pay_label,
                    len({r["agent_name"] for r in table1_rows}),
                    len(table2_rows),
                    _fmt_money(total_anp),
                ]
            )

            prefix = f"anp_commission_{calendar_label}_{stamp}"
            meta = {
                "payout_period": label,
                "invoice_calendar_month": calendar_label,
                "commission_payout_month": pay_label,
                "invoice_date_from": ps.isoformat(),
                "invoice_date_to": pe.isoformat(),
                "agent_types": agent_types_str,
                "total_qualifying_agents": len({r["agent_name"] for r in table1_rows}),
                "total_qualifying_invoices": len(table2_rows),
                "filter_description": (
                    "2026 invoices, 1st payment secured, agent types: internal + FULL TIME"
                ),
                "tier_basis": "Accumulated total amount (invoice total_amount)",
                "anp_timing": "Commission paid in month after invoice_date month",
            }

            dc = out_dir / f"{prefix}_table2_by_customer.csv"
            sc = out_dir / f"{prefix}_table1_accumulated_anp.csv"
            write_csv(dc, table2_rows, TABLE2_CSV_FIELDS)
            write_csv(sc, table1_rows, TABLE1_CSV_FIELDS)

            xlsx_path = out_dir / f"{prefix}.xlsx"
            if not args.no_excel:
                try:
                    write_excel(xlsx_path, table1_rows, table2_rows, meta)
                except ImportError:
                    print("openpyxl not installed; skipped Excel.", file=sys.stderr)

            if args.print_each_month:
                print(f"\n--- {calendar_label} (paid {pay_label}) ---")
                try:
                    detail_limit = (
                        None if args.full_console else max(0, args.console_detail_limit)
                    )
                    print_report_tables(meta, table1_rows, table2_rows, detail_limit)
                except ImportError:
                    print(
                        "Install tabulate for table output: pip install tabulate",
                        file=sys.stderr,
                    )

        if not overview_rows:
            print(
                f"No qualifying invoices found for invoicing months in {year} "
                "(check agent types and filters).",
                file=sys.stderr,
            )
            return 0

        print()
        print("=" * 100)
        print(f"ANP COMMISSION - ALL INVOICE MONTHS IN {year}")
        print("=" * 100)
        print(
            "One report per invoicing calendar month where data exists "
            "(first payment secured). Commission is paid in the following month.",
        )
        print()
        try:
            print_year_month_overview(overview_rows)
        except ImportError:
            for row in overview_rows:
                print("  ", row)
        print()

        overview_csv = (
            out_dir / f"anp_commission_{year}_months_overview_{stamp}.csv"
        )
        overview_csv.parent.mkdir(parents=True, exist_ok=True)
        with overview_csv.open("w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh)
            writer.writerow(
                [
                    "invoice_month",
                    "commission_payout_month",
                    "agent_count",
                    "invoice_count",
                    "total_anp_rm",
                ]
            )
            for row in overview_rows:
                writer.writerow(row)

        print(f"Month overview CSV: {overview_csv.resolve()}")
        print(f"Generated {len(overview_rows)} monthly report bundles under: {out_dir.resolve()}")

        return 0

    period_start: date | None = None
    period_end: date | None = None

    if args.all_time:
        payout_label = "all-time"
        invoice_calendar_month = "all"
        commission_payout_month = "N/A"
        if args.payout_month:
            print(
                "Warning: --payout-month ignored when --all-time is set.", file=sys.stderr
            )
    elif default_invoice_year_mode:
        period_start = date(2026, 1, 1)
        period_end = date(2026, 12, 31)
        invoice_calendar_month = "2026-01..2026-12"
        commission_payout_month = "varies by invoice month"
        payout_label = "invoice-year-2026"
    else:
        if args.payout_month:
            py, pm = map(int, args.payout_month.split("-"))
        else:
            tdy = date.today()
            py, pm = tdy.year, tdy.month
        period_start, period_end = invoice_period_bounds(py, pm)
        invoice_calendar_month = (
            f"{period_start.year:04d}-{period_start.month:02d}"
            if period_start
            else "unknown"
        )
        commission_payout_month = f"{py:04d}-{pm:02d}"
        payout_label = f"inv-{invoice_calendar_month}_pay-{commission_payout_month}"

    table1_rows, table2_rows = build_anp_tables(
        agents, invoices, customers, period_start, period_end
    )

    report_year = period_start.year if period_start else date.today().year
    total_users = count_distinct_agents_in_year(
        invoices, set(agent_ids), report_year
    )

    prefix = f"anp_commission_{payout_label}_{stamp}"
    table1_csv = out_dir / f"{prefix}_table1_accumulated_anp.csv"
    table2_csv = out_dir / f"{prefix}_table2_by_customer.csv"
    write_csv(table1_csv, table1_rows, TABLE1_CSV_FIELDS)
    write_csv(table2_csv, table2_rows, TABLE2_CSV_FIELDS)

    meta = {
        "payout_period": payout_label,
        "invoice_calendar_month": invoice_calendar_month,
        "commission_payout_month": commission_payout_month,
        "invoice_date_from": period_start.isoformat() if period_start else "all",
        "invoice_date_to": period_end.isoformat() if period_end else "all",
        "agent_types": agent_types_str,
        "total_qualifying_agents": total_users,
        "total_qualifying_invoices": len(table2_rows),
        "filter_description": (
            "2026 invoices, 1st payment secured, agent types: internal + FULL TIME"
        ),
        "tier_basis": "Accumulated total amount (invoice total_amount)",
        "anp_timing": "Commission paid in month after invoice_date month",
    }

    xlsx_path = out_dir / f"{prefix}.xlsx"
    if not args.no_excel:
        try:
            write_excel(xlsx_path, table1_rows, table2_rows, meta)
        except ImportError:
            print("openpyxl not installed; skipped Excel.", file=sys.stderr)

    try:
        detail_limit = None if args.full_console else max(0, args.console_detail_limit)
        print_report_tables(meta, table1_rows, table2_rows, detail_limit)
    except ImportError:
        print("Install tabulate for table output: pip install tabulate", file=sys.stderr)

    print("Files saved:")
    print(f"  Excel (tables): {xlsx_path.resolve() if xlsx_path.exists() else '(skipped)'}")
    print(f"  Table 1 CSV:    {table1_csv.resolve()}")
    print(f"  Table 2 CSV:    {table2_csv.resolve()}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
